#!/usr/bin/env python3
"""
Level 2: Excel price sync + marketplace XML export.

- Updates Import prices from Price list
- Sets old_price = price + 10%
- Highlights changed / identical rows
- Streams XML in chunks for 10k+ products
"""

from __future__ import annotations

import argparse
import xml.sax.saxutils as saxutils
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from price_loader import load_price_map

CHUNK_SIZE = 500
OLD_PRICE_MULTIPLIER = 1.10
HIGHLIGHT_COLUMNS = 12

FILL_CHANGED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_IDENTICAL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

SKU_KEYS = ("артикул", "sku", "id", "код", "article")
PRICE_KEYS = ("ціна", "цена", "price", "cost")
NAME_KEYS = ("назва", "название", "name", "title", "товар")
OLD_PRICE_KEYS = ("old_price", "oldprice", "стара ціна", "старая цена", "стара_ціна")
CATEGORY_KEYS = ("категорія", "категория", "category", "cat", "родитель")
PICTURE_KEYS = ("галерея", "gallery", "picture", "image", "фото")


@dataclass
class ProductRow:
    offer_id: str
    sku: str
    name: str
    price: float
    old_price: float
    category: str
    picture: str
    status: str


def normalize(value: object) -> str:
    return str(value or "").strip()


def normalize_key(value: object) -> str:
    return normalize(value).lower().replace("_", " ")


def find_column_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    normalized = [normalize_key(header) for header in headers]
    for candidate in candidates:
        if candidate in normalized:
            return normalized.index(candidate)
    return None


def parse_price(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    cleaned = normalize(value).replace(" ", "").replace(",", ".")
    return round(float(cleaned), 2)


def get_import_sheet_name(import_path: Path) -> str | None:
    workbook = load_workbook(import_path, read_only=True)
    sheet_name = None
    for name in workbook.sheetnames:
        if "export" in name.lower() or "імпорт" in name.lower() or "import" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = workbook.sheetnames[0]
    workbook.close()
    return sheet_name


def resolve_new_price(sku: str, original_price: float, price_map: dict[str, float]) -> float:
    if sku in price_map:
        return price_map[sku]
    upper = sku.upper()
    for key, value in price_map.items():
        if key.upper() == upper:
            return value
    return original_price


def iter_import_rows(
    import_path: Path,
    price_map: dict[str, float],
    sheet_name: str | None = None,
) -> Iterator[ProductRow]:
    workbook = load_workbook(import_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name or workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize(value) for value in next(rows)]

    id_idx = find_column_index(headers, ("id",))
    sku_idx = find_column_index(headers, SKU_KEYS)
    name_idx = find_column_index(headers, NAME_KEYS)
    price_idx = find_column_index(headers, PRICE_KEYS)
    category_idx = find_column_index(headers, CATEGORY_KEYS)
    picture_idx = find_column_index(headers, PICTURE_KEYS)

    if sku_idx is None or price_idx is None:
        workbook.close()
        raise ValueError("Import file must contain article and price columns.")

    for row in rows:
        if not row or row[sku_idx] in (None, ""):
            continue

        sku = normalize(row[sku_idx])
        original_price = parse_price(row[price_idx])
        new_price = resolve_new_price(sku, original_price, price_map)
        old_price = round(new_price * OLD_PRICE_MULTIPLIER, 2)
        status = "identical" if abs(new_price - original_price) < 0.001 else "changed"
        picture = ""
        if picture_idx is not None and row[picture_idx]:
            picture = normalize(row[picture_idx]).split("|")[0]

        yield ProductRow(
            offer_id=normalize(row[id_idx]) if id_idx is not None and row[id_idx] is not None else sku,
            sku=sku,
            name=normalize(row[name_idx]) if name_idx is not None else sku,
            price=new_price,
            old_price=old_price,
            category=normalize(row[category_idx]) if category_idx is not None else "",
            picture=picture,
            status=status,
        )

    workbook.close()


def write_updated_excel(
    import_path: Path,
    output_path: Path,
    price_map: dict[str, float],
    sheet_name: str | None = None,
) -> tuple[int, int]:
    source = load_workbook(import_path)
    sheet: Worksheet = source[sheet_name] if sheet_name else source.active
    headers = [normalize(cell.value) for cell in sheet[1]]

    sku_idx = find_column_index(headers, SKU_KEYS)
    price_idx = find_column_index(headers, PRICE_KEYS)
    old_price_idx = find_column_index(headers, OLD_PRICE_KEYS)

    if sku_idx is None or price_idx is None:
        raise ValueError("Import file must contain article and price columns.")

    if old_price_idx is None:
        old_price_idx = len(headers)
        sheet.cell(row=1, column=old_price_idx + 1, value="Старая цена")
        headers.append("Старая цена")

    changed_count = 0
    identical_count = 0
    highlight_cols = min(HIGHLIGHT_COLUMNS, len(headers))

    for row_idx in range(2, sheet.max_row + 1):
        sku = normalize(sheet.cell(row=row_idx, column=sku_idx + 1).value)
        if not sku:
            continue

        original_price = parse_price(sheet.cell(row=row_idx, column=price_idx + 1).value)
        new_price = resolve_new_price(sku, original_price, price_map)
        old_price = round(new_price * OLD_PRICE_MULTIPLIER, 2)

        sheet.cell(row=row_idx, column=price_idx + 1, value=new_price)
        sheet.cell(row=row_idx, column=old_price_idx + 1, value=old_price)

        if abs(new_price - original_price) < 0.001:
            identical_count += 1
            fill = FILL_IDENTICAL
        else:
            changed_count += 1
            fill = FILL_CHANGED

        for col_idx in range(1, highlight_cols + 1):
            sheet.cell(row=row_idx, column=col_idx).fill = fill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    source.save(output_path)
    return changed_count, identical_count


def xml_escape(value: str) -> str:
    return saxutils.escape(value)


def write_offer_xml(handle, product: ProductRow) -> None:
    handle.write(f'    <offer id="{xml_escape(product.offer_id)}" available="true">\n')
    handle.write(f"      <vendorCode>{xml_escape(product.sku)}</vendorCode>\n")
    handle.write(f"      <name>{xml_escape(product.name)}</name>\n")
    handle.write(f"      <price>{product.price:.2f}</price>\n")
    handle.write(f"      <oldprice>{product.old_price:.2f}</oldprice>\n")
    if product.category:
        handle.write(f"      <categoryId>{xml_escape(product.category)}</categoryId>\n")
    if product.picture:
        handle.write(f"      <picture>{xml_escape(product.picture)}</picture>\n")
    handle.write("    </offer>\n")


def write_marketplace_xml(
    products: Iterator[ProductRow],
    output_path: Path,
    chunk_size: int = CHUNK_SIZE,
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    buffer: list[ProductRow] = []
    catalog_date = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")

    with output_path.open("w", encoding="utf-8") as handle:
        handle.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        handle.write(f'<yml_catalog date="{catalog_date}">\n')
        handle.write("  <shop>\n")
        handle.write("    <name>Ecopol Catalog</name>\n")
        handle.write("    <company>Developer Test</company>\n")
        handle.write("    <url>https://ecopol.com.ua</url>\n")
        handle.write("    <currencies>\n")
        handle.write('      <currency id="UAH" rate="1"/>\n')
        handle.write("    </currencies>\n")
        handle.write("    <offers>\n")

        for product in products:
            buffer.append(product)
            total += 1
            if len(buffer) >= chunk_size:
                for item in buffer:
                    write_offer_xml(handle, item)
                buffer.clear()

        for item in buffer:
            write_offer_xml(handle, item)

        handle.write("    </offers>\n")
        handle.write("  </shop>\n")
        handle.write("</yml_catalog>\n")

    return total


def process(import_path: Path, price_path: Path, output_dir: Path, chunk_size: int) -> None:
    sheet_name = get_import_sheet_name(import_path)
    price_map = load_price_map(price_path)
    excel_output = output_dir / "Import_оновлений.xlsx"
    xml_output = output_dir / "catalog.xml"

    changed, identical = write_updated_excel(
        import_path, excel_output, price_map, sheet_name=sheet_name
    )
    product_iter = iter_import_rows(import_path, price_map, sheet_name=sheet_name)
    total = write_marketplace_xml(product_iter, xml_output, chunk_size=chunk_size)

    print(f"Price entries loaded: {len(price_map)}")
    print(f"Excel: {excel_output}")
    print(f"XML:   {xml_output}")
    print(f"Rows processed: {total}")
    print(f"Changed: {changed}, Identical: {identical}")


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Process Import/Price Excel files.")
    parser.add_argument(
        "--import",
        dest="import_file",
        default=base_dir / "data" / "Import.xlsx",
        type=Path,
        help="Path to Import Excel file",
    )
    parser.add_argument(
        "--price",
        dest="price_file",
        default=base_dir / "data" / "Прайс.xlsx",
        type=Path,
        help="Path to Price Excel file",
    )
    parser.add_argument(
        "--output-dir",
        default=base_dir / "output",
        type=Path,
        help="Directory for output files",
    )
    parser.add_argument(
        "--chunk-size",
        default=CHUNK_SIZE,
        type=int,
        help="XML offers per write chunk",
    )
    args = parser.parse_args()

    if not args.import_file.exists():
        raise FileNotFoundError(f"Import file not found: {args.import_file}")
    if not args.price_file.exists():
        raise FileNotFoundError(f"Price file not found: {args.price_file}")

    process(args.import_file, args.price_file, args.output_dir, args.chunk_size)


if __name__ == "__main__":
    main()
