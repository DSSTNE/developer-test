"""Load prices from multi-sheet supplier price list."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

PRICE_HINTS = ("цена", "ціна", "грн", "price", "ррц", "роздріб", "розница", "rrc")
ART_HINTS = ("артикул", "article", "sku", "код")
SKIP_SHEETS = {"головна"}


def normalize_article(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def normalize_header(value: object) -> str:
    return normalize_article(value).lower().replace("_", " ")


def parse_price(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    cleaned = normalize_article(value).replace(" ", "").replace("\xa0", "")
    match = re.search(r"([\d]+(?:[.,]\d+)?)", cleaned)
    if not match:
        return None
    return round(float(match.group(1).replace(",", ".")), 2)


def format_area(area: float) -> str:
    if abs(area - round(area)) < 0.001:
        return str(int(round(area)))
    text = f"{area:.1f}".rstrip("0").rstrip(".")
    return text


def is_price_header(value: object) -> bool:
    text = normalize_header(value).replace("\n", " ")
    return any(hint in text for hint in PRICE_HINTS)


def is_article_header(value: object) -> bool:
    text = normalize_header(value)
    return any(hint in text for hint in ART_HINTS)


def add_price(price_map: dict[str, float], article: str, price: float | None) -> None:
    if not article or price is None or price <= 0:
        return
    price_map[article] = price


def extract_extherm(row: tuple) -> list[tuple[str, float]]:
    if not row or not row[0]:
        return []
    name = str(row[0])
    price = None
    for cell in reversed(row):
        parsed = parse_price(cell)
        if parsed is not None:
            price = parsed
            break
    if price is None:
        return []

    results: list[tuple[str, float]] = []
    patterns = [
        (r"ET ECO (\d+)-180", lambda m: f"ET-ECO-{m.group(1)}"),
        (r"ETС ECO 20-(\d+)", lambda m: f"ETC-ECO-20-{m.group(1)}"),
        (r"ETC ECO 20-(\d+)", lambda m: f"ETC-ECO-20-{m.group(1)}"),
    ]
    for pattern, builder in patterns:
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            results.append((builder(match), price))
    return results


def extract_devi(row: tuple) -> list[tuple[str, float]]:
    if not row or not row[0]:
        return []
    name = str(row[0])
    model_match = re.search(r"\((DTIR-\d+|DTIP-\d+)\)", name, re.IGNORECASE)
    if not model_match:
        return []

    area = None
    for cell in row[1:4]:
        if isinstance(cell, (int, float)) and 0 < float(cell) <= 100:
            area = float(cell)
            break
    price = parse_price(row[3]) if len(row) > 3 else None
    if area is None or price is None:
        return []

    model = model_match.group(1).upper()
    return [(f"{model}-{format_area(area)}", price)]


def extract_fh_series(row: tuple) -> list[tuple[str, float]]:
    if len(row) < 2 or not row[1]:
        return []
    name = normalize_article(row[1])
    price = parse_price(row[5]) if len(row) > 5 else None
    if price is None:
        price = parse_price(row[4]) if len(row) > 4 else None
    if price is None:
        return []

    results: list[tuple[str, float]] = []
    match = re.match(r"FH L (\d+)", name, re.IGNORECASE)
    if match:
        suffix = match.group(1)
        area = int(suffix) / 10
        results.append((f"FH-EC-{format_area(area)}", price))
        return results

    match = re.match(r"FH P (\d+)", name, re.IGNORECASE)
    if match:
        suffix = match.group(1)
        area = int(suffix) / 10
        results.append((f"FHP-{format_area(area)}", price))
    return results


def extract_direct_article_row(row: tuple) -> list[tuple[str, float]]:
    if not row:
        return []

    article = normalize_article(row[0])
    if not article or len(article) < 2:
        return []

    if not re.match(r"^[A-Za-zА-Яа-яІіЇїЄєҐґ0-9][A-Za-zА-Яа-яІіЇїЄєҐґ0-9\-/. ]+$", article):
        return []

    prices = [parse_price(cell) for cell in row[1:8]]
    prices = [p for p in prices if p is not None]
    if not prices:
        return []

    return [(article.strip(), prices[-1])]


def extract_ryxon_hm(row: tuple) -> list[tuple[str, float]]:
    if not row or not row[0]:
        return []
    article = normalize_article(row[0]).replace(" ", "")
    match = re.match(r"HM-200-([\d.]+)", article, re.IGNORECASE)
    if not match:
        return []
    price = parse_price(row[5]) if len(row) > 5 and parse_price(row[5]) else parse_price(row[4])
    if price is None:
        return []
    area = match.group(1).rstrip("0").rstrip(".") if "." in match.group(1) else match.group(1)
    return [(f"Ryxon-HM-{area}", price)]


def extract_woks_mappings(rows: list[tuple]) -> dict[str, float]:
    mapping: dict[str, float] = {}
    for row in rows:
        if not row or not row[1]:
            continue
        text = str(row[1])
        price = parse_price(row[2]) if len(row) > 2 else None
        if price is None:
            continue

        length_match = re.search(r"\((\d+(?:[.,]\d+)?)\s*м\)", text, re.IGNORECASE)
        if not length_match:
            continue
        length = length_match.group(1).replace(",", ".")

        if "Woks-10" in text:
            mapping[f"WOKS-10-{length}"] = price
        elif "Woks-18" in text:
            mapping[f"woks17k-{length}"] = price
        elif "WoksMat" in text:
            area_match = re.search(r"\(([\d,]+)\s*м2\)", text, re.IGNORECASE)
            if area_match:
                area = area_match.group(1).replace(",", ".")
                mapping[f"WOKSM-{area}"] = price

    return mapping


def parse_table_section(
    rows: list[tuple],
    start_index: int,
    price_map: dict[str, float],
) -> int:
    header = rows[start_index]
    art_cols = [i for i, value in enumerate(header) if is_article_header(value)]
    price_cols = [i for i, value in enumerate(header) if is_price_header(value)]
    if not art_cols or not price_cols:
        return start_index

    art_idx = art_cols[0]
    price_idx = price_cols[-1]
    empty_streak = 0

    for row in rows[start_index + 1 :]:
        if not row or all(cell in (None, "") for cell in row):
            empty_streak += 1
            if empty_streak > 20:
                break
            continue
        empty_streak = 0

        if is_article_header(row[art_idx] if art_idx < len(row) else None):
            break

        article = normalize_article(row[art_idx] if art_idx < len(row) else None)
        price = parse_price(row[price_idx] if price_idx < len(row) else None)
        add_price(price_map, article, price)

        for article, parsed_price in extract_ryxon_hm(row):
            add_price(price_map, article, parsed_price)

    return start_index + 1


def load_price_map(price_path: Path) -> dict[str, float]:
    workbook = load_workbook(price_path, read_only=True, data_only=True)
    price_map: dict[str, float] = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if sheet_name == "Woks":
            price_map.update(extract_woks_mappings(rows))

        for index, row in enumerate(rows):
            if row and any(is_article_header(cell) for cell in row if cell is not None):
                parse_table_section(rows, index, price_map)

            for article, price in extract_extherm(row):
                add_price(price_map, article, price)
            for article, price in extract_devi(row):
                add_price(price_map, article, price)
            for article, price in extract_fh_series(row):
                add_price(price_map, article, price)
            for article, price in extract_direct_article_row(row):
                add_price(price_map, article, price)
            for article, price in extract_ryxon_hm(row):
                add_price(price_map, article, price)

    workbook.close()
    return price_map
